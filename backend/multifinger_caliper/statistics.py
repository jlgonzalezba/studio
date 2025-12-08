
from joints import data, dept, collars
import pandas as pd
import numpy as np
import os
from pipes_list import pipes_data
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles import Font, PatternFill







#LO QUE SE IMPORTA

#DATA VIENE CON DEPTH Y CON LOS RADIOS CENTRALIZADOS
#DEPT, SOLO LA COLUMNA DE DEPTH
#COLLARS, TABLA DE COLLARS


#LO QUE SE CREA

#DF, DATA DE DIAMETROS, SIN DEPTH
#PIPES, TABLA DE DEPTH INICIO, INICIO Y DATA DE DIAMETROS DE CADA TUBO
#SIZE, HACE PARTE DE LA TABLA DE INTEGRIDAD FINA, TIENE DEPTH INICIO, FIN, LENGTH , OD, ID, PESO



# constantes
tally_difference = 0.6/100


collars = collars.to_numpy()
#tablas de tubulares 
data_fingers = data[1:,1:]  # se elimima la columna de profundidad y se dejan solo los fingers para  posteriormente pasar de radios a diametros


# funcion para convertir radios a diametros y calcular min. max, mean, ademas de transformar la data resultante a pandas
def data_and_min_max_mean(data_fingers,dept):
    N = data_fingers.shape[1] # numero de fingers
    half = N//2  # mitad de numero de fingers

    diameters_full = np.zeros((data_fingers.shape[0], N)) # se crea un np con ceros con las dimensiones de # de filas cy # de columnas / 2
    diameters_half = np.zeros((data_fingers.shape[0], half))

    # se llena el array con los diametros
    for i in range(half):
        diameters_half[:,i] = data_fingers[:,i] + data_fingers[:, i+half]
        
    for i in range(N):
        diameters_full[:,i] = data_fingers[:,i]*2


    data_fingers = diameters_full.copy()  # ahora fingers tiene la data de diametros

    mean_col = diameters_full.mean(axis=1)
    min_col = diameters_half.min(axis=1)
    max_col = diameters_full.max(axis=1)

    mean_col = mean_col[:,None]
    min_col = min_col[:,None]
    max_col = max_col[:,None]
    dept_col=dept[:,None]
    data_fingers = np.hstack([dept_col,data_fingers, min_col, max_col, mean_col])

    # ahora se convertira todo a pandas

    cols_diam = [f"D{i+1:02d}" for i in range(N)]
    columns = ["DEPT"] + cols_diam + ["MIN","MAX","MEAN"]
    
    data_fingers = pd.DataFrame(data_fingers, columns=columns) 

    return data_fingers

df = data_and_min_max_mean(data_fingers,dept)



# Crear una lista con la data de cada uno de los tubos detectados
pipes = []
for i in range(len(collars)-1):

    idx1 = df["DEPT"].searchsorted(collars[i,1])
    idx2 = df["DEPT"].searchsorted(collars[i+1,0])
    new_pipe = df[idx1+1:idx2]
    new_pipe = pd.DataFrame(new_pipe)
    pipes.append(new_pipe)



# calcular OD, ID, Weight, Top, Bottom, Length

def start_pipe_statistics(pipes):

    #calcular tamaño del casing

    size = pd.DataFrame(columns=["TOP","BOTTOM","LENGTH","OD","ID","WEIGHT"])
    desviation=[]
    for pipe in pipes:
        new_desviation = pipe["MEAN"].var(ddof=0)*1000 # calcula la varianza de la columna MEAN
        desviation.append(new_desviation) 
        mean = pipe["MEAN"].mean()

        if new_desviation < 1.5:   # Solo asginar tañano de casing a los tubos donde mean no tiene mucha varianza
            new_size_down = pipes_data.loc[pipes_data["ID"] <= mean].nlargest(1, "ID")
            new_size_up = pipes_data.loc[pipes_data["ID"] >= mean].nsmallest(1, "ID")


            if new_size_down.empty or new_size_up.empty:
                continue
            
            #difference_dowm = abs(mean - new_size_down["ID"].iloc[0])
            difference_up = abs(mean - new_size_up["ID"].iloc[0])

            # Normalmente asignar el tamaño de casing donde el ID nominal esta por debajo de mean....
            # .... pero si el ID nominal proxima hacia arriba esta muy cerca, se lo asigna.

            if difference_up < mean*tally_difference:  #tally_difference arriba en constantes
                od = new_size_up["OD"].iloc[0]
                id = new_size_up["ID"].iloc[0]
                weigth = new_size_up["WEIGHT"].iloc[0]
                size.loc[len(size)]= {"TOP":pipe.iloc[0,0],"BOTTOM":pipe.iloc[-1,0],"LENGTH":pipe.iloc[-1,0]-pipe.iloc[0,0],"OD":od, "ID":id, "WEIGHT":weigth}
                
            else:
                od = new_size_down["OD"].iloc[0]
                id = new_size_down["ID"].iloc[0]
                weigth = new_size_down["WEIGHT"].iloc[0]
                size.loc[len(size)]= {"TOP":pipe.iloc[0,0],"BOTTOM":pipe.iloc[-1,0],"LENGTH":pipe.iloc[-1,0]-pipe.iloc[0,0],"OD":od, "ID":id, "WEIGHT":weigth}

        else:
            size.loc[len(size)]= {"TOP":pipe.iloc[0,0],"BOTTOM":pipe.iloc[-1,0],"LENGTH":pipe.iloc[-1,0]-pipe.iloc[0,0],"OD":0, "ID":0, "WEIGHT":0}    
        
        
        def local_mode(size: pd.Series, pos: int, k: int=5, cols=None):

            if cols is None:
                cols = ["OD", "ID", "WEIGHT"]

            n = len(size)
            start = max(0, pos -k)
            end = min(n-1, pos +k)

            window = size.iloc[start:end+1]
            mode = window[cols].mode().iloc[0]
            return mode

        idxs = size.index[size["OD"] == 0].to_list()  # lista de indices donde OD = 0

        for idx in idxs:
            mode = local_mode(size, pos=idx, k=5, cols=["OD","ID","WEIGHT"])
            size.loc[idx,["OD","ID","WEIGHT"]] = mode
        
    # agregar columna item

    size.insert(0,"ITEM", range(1,len(size)+1))
       
    return size


statistics_table = start_pipe_statistics(pipes)


def continue_statistics(statistics_table, pipes):

    #statistics_table["MIN_ID","MIN_ID_D","MAX_LOSS","MAX_LOSS_D,","MAX_PEN","MAX_PEN_D","MAX_PEN_PERC"]=pd.NA
    
    for i in range(len(statistics_table)):
        pipe = pipes[i]

        #Diametro minimo y depth de diametro minimo
        idx = pipe["MIN"].idxmin()
        min_id = round(pipe.loc[idx,"MIN"],3)
        min_id_d = pipe.loc[idx,"DEPT"] 

        #Diametro maxima penetracion, prof maxima pen, % max pen
        idx = pipe["MAX"].idxmax()
        max_id = round(pipe.loc[idx,"MAX"],3)
        max_id_d = pipe.loc[idx,"DEPT"] 

        od_id = statistics_table.loc[i, "OD"]  - statistics_table.loc[i,"ID"]
        id = statistics_table.loc[i,"ID"]
        penetration = round((max_id-id)/(od_id)*100,2)
        
        if penetration < 0:
            penetration = 0
        elif penetration > 100:
            penetration = 100

        if penetration<=20:
            clasification = 'I'
        elif 20<penetration<=40:
            clasification = 'II'
        elif 40<penetration<=60:
            clasification = 'III'
        elif 60<penetration<=80:
            clasification = 'IV'
        elif penetration > 80:
            clasification = 'V'


        


        #insetar valores a statistics_table
        statistics_table.loc[i,"MIN_ID"]=min_id
        statistics_table.loc[i,"MIN_ID_DEPTH"]=min_id_d

        statistics_table.loc[i,"MAX_ID"]=max_id
        statistics_table.loc[i,"MAX_ID_DEPTH"]=max_id_d
        statistics_table.loc[i,"MAX_PEN_%"]=penetration
        statistics_table.loc[i,"CLASS_PEN"]=clasification

    def color_cell(clasification):
        return(
            "background-color:yellow;" if clasification == 'I' else
            "background-color:green;" if clasification == 'II' else
            "background-color:orange;" if clasification == 'III' else
            "background-color:orangered;" if clasification == 'IV' else
            "background-color:darkred;" if clasification == 'V' else
            ""
        )
        
    styled = statistics_table.style.map(color_cell, subset=pd.IndexSlice[:, "CLASS_PEN"])
    
    return statistics_table, styled




statistics_table,styled = continue_statistics(statistics_table,pipes)









######STIYE EL EXCEL

export_dir = os.path.join(os.path.dirname(__file__), '..', '..', 'exports')
styled.to_excel(os.path.join(export_dir,"statistics_format.xlsx"),index=False)
# Load the exported Excel file
wb = load_workbook(os.path.join(export_dir, "statistics_format.xlsx"))
ws = wb.active
# Set row heights for specific rows (adjust row numbers and heights as needed)
ws.row_dimensions[1].height = 30  # Row 1

# Save the modified file
for row in [1]:  # The rows you modified
    for col in range(1, ws.max_column + 1):  # All columns in the row
        cell = ws.cell(row=row, column=col)
        cell.alignment = Alignment(wrap_text=True)


for col in range(1, ws.max_column + 1):  # All columns in the row
    cell = ws.cell(row=1, column=col)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


# Center text in an entire column (e.g., column 'A' for all rows)
  # Change to the desired column letter (e.g., 'B', 'C')
for row in range(1, ws.max_row + 1):
    cell = ws.cell(row=row, column=13)  # Get column number from letter
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Apply borders to all cells with values
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value is not None and str(cell.value).strip() != '':
            cell.border = thin_border



# Cambiar color de fondo y fuente de la fila 1

row_num = 1  # Change to the desired row number
font_color = 'FFFFFF'  # font 
fill_color = '708090'  # fill 
for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=row_num, column=col)
    cell.font = Font(color=font_color, bold=True)
    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

# Save the modified file

wb.save(os.path.join(export_dir, "statistics_format.xlsx"))

statistics_table.to_csv(os.path.join(export_dir, 'table_statistics.csv'), index=False, float_format='%.4f')




### FIN STYLE



data= pd.DataFrame(data)
data.to_csv(os.path.join(os.path.dirname(__file__), '..', '..', 'exports', 'data.csv'), index=False, float_format='%.4f')





#CREAR PANDAS CON LOS OD's & ID's Y SU PROFUNDIDAD

def od_id(dept, statistics_table):

    df_od_id = pd.DataFrame(columns = ["DEPT","OD","ID"])

    dept_col=dept[:,None]
    tops = statistics_table["BOTTOM"].to_numpy()
    for dept in dept_col:


        idx = np.searchsorted(tops, dept)
        idx=int(idx)
        if idx >=len(tops):
            idx=len(tops)-1

        df_od_id.loc[len(df_od_id)] = {
            "DEPT":float(dept), 
            "OD":statistics_table.iat[idx,statistics_table.columns.get_loc("OD")],
            "ID":statistics_table.iat[idx,statistics_table.columns.get_loc("ID")]
        }
    return df_od_id


df_id_od = od_id(dept,statistics_table)
df_id_od.to_csv(os.path.join(os.path.dirname(__file__), '..', '..', 'exports', 'OD_ID.csv'), index=False, float_format='%.4f')